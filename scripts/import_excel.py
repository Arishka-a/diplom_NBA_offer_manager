# -*- coding: utf-8 -*-
"""
Импорт реальных данных из NBA_offers.xlsx в БД nba_offermanager.

Что делает:
  1. Удаляет существующие офферы и связи offer_segments (демо-данные).
     Сегменты, клиенты, пользователи — НЕ трогает.
  2. Импортирует products (19), vas (55), availability_rules (33).
  3. Импортирует 251 реальный оффер с деталями (offer_details) и каналами (offer_channels).
  4. Привязывает каждый оффер к поведенческим сегментам (1-6) по правилам.

Запуск:
    pip install openpyxl psycopg2-binary
    python scripts/import_excel.py

Параметры подключения берутся из переменных окружения с дефолтами под docker-compose.
"""
import os
import sys
from pathlib import Path

# Форсим UTF-8 для psycopg2 — иначе на Windows кириллические сообщения от postgres
# приходят в cp1251 и падают с UnicodeDecodeError при connect().
os.environ['PGCLIENTENCODING'] = 'UTF8'

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

EXCEL_PATH = Path(__file__).resolve().parent.parent / 'NBA_offers.xlsx'

DB_CONFIG = {
    'host': os.getenv('DB_HOST', 'localhost'),
    'port': int(os.getenv('DB_PORT', '5432')),
    'dbname': os.getenv('DB_NAME', 'nba_offermanager'),
    'user': os.getenv('DB_USER', 'nba_user'),
    'password': os.getenv('DB_PASSWORD', 'nbaDB123'),
    'client_encoding': 'UTF8',
    'options': '-c lc_messages=en_US.UTF-8',
}

ADMIN_USER_ID = 1  # admin из 01_init.sql

# Маппинг каналов (колонки Excel → значение для БД)
CHANNEL_COLUMNS = {
    'actual_phone': 'phone',
    'actual_push':  'push',
    'actual_mail':  'mail',
    'actual_sms':   'sms',
    'actual_lm':    'lm',
    'actual_okc':   'okc',
    'actual_ltv':   'ltv',
}


# ─────────────────────────────────────────────────────────────
# Утилиты чтения Excel
# ─────────────────────────────────────────────────────────────

def sheet_to_dicts(ws):
    """Превращает лист в список словарей (по заголовкам в строке 1).
    Пропускает строки, где первая ячейка пустая."""
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value is None:
            continue
        row = {}
        for i, h in enumerate(headers, start=1):
            if h is None:
                continue
            row[h] = ws.cell(row=r, column=i).value
        rows.append(row)
    return rows


def to_int(v):
    if v is None or v == '':
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def to_float(v):
    if v is None or v == '':
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def to_str(v, max_len=None):
    if v is None:
        return None
    s = str(v).strip()
    if not s or s == '-':
        return None
    if max_len and len(s) > max_len:
        s = s[:max_len]
    return s


# ─────────────────────────────────────────────────────────────
# Классификация офферов
# ─────────────────────────────────────────────────────────────

def classify_offer_type(row):
    """Решает, какой offer_type (из существующего enum) присвоить.
    Допустимые значения: discount, bonus, recommendation, upgrade, retention."""
    product = (to_str(row.get('PRODUCT')) or '').upper()
    offer   = (to_str(row.get('OFFER')) or '').lower()
    target  = (to_str(row.get('segment_name')) or '').lower()

    discount_pct = to_float(row.get('internet_tv_discount_pct')) or 0
    discount_abs = to_float(row.get('internet_tv_discount_abs')) or 0
    hw_int_disc  = to_float(row.get('hardware_internet_discount_pct')) or 0
    hw_tv_disc   = to_float(row.get('hardware_tv_discount_pct')) or 0
    vas_disc     = to_float(row.get('vas_discount_per')) or 0

    # 1. Retention: промо-офферы для удержания (razgonis = "разгонись")
    if 'razgonis' in target or 'razgonis' in offer or 'retention' in offer:
        return 'retention'

    # 2. Discount: явная скидка
    if discount_pct or discount_abs or hw_int_disc or hw_tv_disc or vas_disc:
        return 'discount'

    # 3. Bonus: рассрочка (installments) или продление VAS
    if 'installments' in offer or row.get('vas_prolongation'):
        return 'bonus'

    # 4. Upgrade: улучшение тарифа
    if product == 'STO_PLUS_PLUS' or 'mega' in target or 'giga_' in offer or product == 'STO_PLUS':
        return 'upgrade'

    # 5. Recommendation: остальное (SVOD, антивирусы, контент)
    return 'recommendation'


def classify_product_category(code):
    """Категория для каталога products."""
    code_u = code.upper()
    if code_u.startswith('SVOD_') or code_u in ('IVI', 'LITRES'):
        return 'svod'
    if code_u in ('ROUTER', 'TV_BOX', 'TV_BOX_CIFR_DOPAKET', 'YASTATION'):
        return 'hardware'
    if code_u in ('STO_PLUS', 'STO_PLUS_PLUS', 'STO_PLUS_ROUTER'):
        return 'tariff'
    return 'service'


def calculate_priority(row):
    """Считает priority (1-100) на основе количества активных каналов и наличия скидок."""
    base = 50
    channels = sum(1 for col in CHANNEL_COLUMNS if (to_float(row.get(col)) or 0) >= 1)
    base += channels * 5  # +5 за каждый активный канал (макс +35)

    if to_float(row.get('internet_tv_discount_pct')) or 0 >= 20:
        base += 5

    return max(1, min(100, base))


def link_to_behavioral_segments(row, offer_type):
    """Возвращает список ID поведенческих сегментов (1-6), к которым привязать оффер."""
    seg_ids = []
    product = (to_str(row.get('PRODUCT')) or '').upper()
    offer   = (to_str(row.get('OFFER')) or '').lower()
    target  = (to_str(row.get('segment_name')) or '').lower()

    # 1: Новые клиенты — базовые офферы и рассрочка
    if 'base_offer' in offer or 'installments' in offer:
        seg_ids.append(1)

    # 2: VIP клиенты — премиум-подписки и топовые тарифы
    if product in ('SVOD_PREMIER', 'SVOD_AMEDIATEKA', 'STO_PLUS_PLUS') \
       or 'giga_500' in target or 'giga_1000' in target or 'giga_1000' in offer:
        seg_ids.append(2)

    # 3: Риск оттока — retention
    if offer_type == 'retention':
        seg_ids.append(3)

    # 4: Потенциал апгрейда
    if offer_type == 'upgrade':
        seg_ids.append(4)

    # 5: Интернет без ТВ — ТВ-продукты и SVOD
    if product in ('TV_BOX', 'TV_BOX_CIFR_DOPAKET', 'MOVIX_ZASMOTRIS') \
       or product.startswith('SVOD_'):
        seg_ids.append(5)

    # 6: Давние клиенты — продление VAS и долгосрочные пакеты
    if row.get('vas_prolongation') or '_36m' in target or 'mono_36m' in target:
        seg_ids.append(6)

    return list(set(seg_ids))


def build_title_and_description(row):
    """Формирует читаемый title (≤100 символов) и description."""
    product = to_str(row.get('PRODUCT')) or 'OFFER'
    offer   = to_str(row.get('OFFER')) or ''
    target  = to_str(row.get('segment_name')) or ''

    title = f"{product} {offer}".strip()
    if target and len(title) + len(target) + 3 <= 95:
        title = f"{title} ({target})"
    title = title[:100]
    if len(title) < 3:
        title = f"Offer #{row.get('economic_id')}"

    # Description: собираем всё полезное
    parts = []
    if row.get('hardware_internet'):
        parts.append(f"Интернет-оборудование: {row['hardware_internet']}"
                     + (f" ({row['hardware_internet_ownership']})" if row.get('hardware_internet_ownership') else ''))
    if row.get('hardware_tv'):
        parts.append(f"ТВ-оборудование: {row['hardware_tv']}"
                     + (f" ({row['hardware_tv_ownership']})" if row.get('hardware_tv_ownership') else ''))
    if row.get('internet_tv_discount_pct'):
        period = row.get('internet_tv_discount_period') or '—'
        parts.append(f"Скидка {row['internet_tv_discount_pct']}% на {period} мес.")
    if row.get('internet_tv_discount_abs'):
        parts.append(f"Скидка {row['internet_tv_discount_abs']} ₽ (абсолютная)")
    if row.get('vas_names'):
        parts.append(f"VAS: {row['vas_names']}")
    if row.get('vas_prolongation'):
        parts.append("Включено продление VAS")
    if row.get('install_price_int'):
        parts.append(f"Стоимость подключения интернета: {row['install_price_int']} ₽")

    description = '. '.join(parts) if parts else f"Оффер по продукту {product}"
    return title, description


# ─────────────────────────────────────────────────────────────
# Импортёры
# ─────────────────────────────────────────────────────────────

def import_products(cur, wb):
    """Импорт уникальных продуктов. Возвращает dict code → product_id."""
    sheet = wb['product_offer']
    rows = sheet_to_dicts(sheet)
    codes = {}
    for row in rows:
        code = to_str(row.get('PRODUCT'))
        if code and code not in codes:
            codes[code] = classify_product_category(code)

    if not codes:
        return {}

    values = [(code, code, cat) for code, cat in codes.items()]
    execute_values(
        cur,
        "INSERT INTO products (code, name, category) VALUES %s "
        "ON CONFLICT (code) DO UPDATE SET category = EXCLUDED.category "
        "RETURNING id, code",
        values
    )
    result = {row[1]: row[0] for row in cur.fetchall()}
    print(f"  ✓ products: {len(result)} записей")
    return result


def import_vas(cur, wb):
    """Импорт VAS-каталога."""
    sheet = wb['vas_list']
    rows = sheet_to_dicts(sheet)

    values = []
    for row in rows:
        name = to_str(row.get('vas_name'))
        if not name:
            continue
        values.append((
            name,
            to_str(row.get('vas_name_calc')),
            to_str(row.get('vas_type_name')),
            to_str(row.get('service_name')),
            to_float(row.get('price_base_vas')),
            to_float(row.get('vas_cost_standalone')),
            to_float(row.get('vas_cost_package')),
        ))

    if values:
        execute_values(
            cur,
            "INSERT INTO vas (name, calc_name, vas_type, service_name, "
            "price_base, cost_standalone, cost_package) VALUES %s",
            values
        )
    print(f"  ✓ vas: {len(values)} записей")


def import_availability_rules(cur, wb):
    """Импорт правил availability из листа avp."""
    sheet = wb['avp']
    rows = sheet_to_dicts(sheet)

    values = []
    for row in rows:
        field = to_str(row.get('FIELD'))
        sql_cond = to_str(row.get('SQL'))
        if not field or not sql_cond:
            continue
        is_actual = bool(to_int(row.get('IS_ACTUAL')))
        values.append((field, sql_cond, is_actual))

    if values:
        execute_values(
            cur,
            "INSERT INTO availability_rules (field_name, sql_condition, is_actual) VALUES %s",
            values
        )
    print(f"  ✓ availability_rules: {len(values)} записей")


def clear_old_offers(cur):
    """Удаляет демо-офферы и их связи. Не трогает сегменты, клиентов, пользователей."""
    cur.execute("DELETE FROM offer_segments")
    cur.execute("DELETE FROM conditions")
    # offer_history ссылается на offers с ON DELETE CASCADE — удалится автоматически
    cur.execute("DELETE FROM offers")
    cur.execute("ALTER SEQUENCE offers_id_seq RESTART WITH 1")
    print("  ✓ старые офферы удалены")


def import_offers(cur, wb, products_map):
    """Главная функция: импорт офферов с деталями, каналами и связями."""
    sheet = wb['offers']
    rows = sheet_to_dicts(sheet)

    inserted_offers = 0
    inserted_details = 0
    inserted_channels = 0
    inserted_links = 0

    for row in rows:
        economic_id = to_int(row.get('economic_id'))
        product_code = to_str(row.get('PRODUCT'))
        if not economic_id or not product_code:
            continue

        title, description = build_title_and_description(row)
        offer_type = classify_offer_type(row)
        priority = calculate_priority(row)
        product_id = products_map.get(product_code)

        # 1. INSERT в offers
        cur.execute(
            """
            INSERT INTO offers (title, description, offer_type, status, priority,
                                start_date, end_date, created_by, product_id, economic_id)
            VALUES (%s, %s, %s, 'active', %s,
                    NOW(), NOW() + INTERVAL '180 days', %s, %s, %s)
            ON CONFLICT (economic_id) DO NOTHING
            RETURNING id
            """,
            (title, description, offer_type, priority,
             ADMIN_USER_ID, product_id, economic_id)
        )
        result = cur.fetchone()
        if not result:
            continue
        offer_id = result[0]
        inserted_offers += 1

        # 2. INSERT в offer_details
        cur.execute(
            """
            INSERT INTO offer_details (
                offer_id, target_code,
                hardware_internet, hardware_internet_ownership,
                hardware_tv, hardware_tv_ownership,
                discount_abs, discount_pct, discount_period_months,
                hardware_internet_discount_pct, hardware_tv_discount_pct,
                vas_names, install_price_int, install_price_tv
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                offer_id,
                to_str(row.get('segment_name'), 200),
                to_str(row.get('hardware_internet'), 200),
                to_str(row.get('hardware_internet_ownership'), 100),
                to_str(row.get('hardware_tv'), 200),
                to_str(row.get('hardware_tv_ownership'), 100),
                to_float(row.get('internet_tv_discount_abs')),
                to_float(row.get('internet_tv_discount_pct')),
                to_int(row.get('internet_tv_discount_period')),
                to_float(row.get('hardware_internet_discount_pct')),
                to_float(row.get('hardware_tv_discount_pct')),
                to_str(row.get('vas_names')),
                to_float(row.get('install_price_int')),
                to_float(row.get('install_price_tv')),
            )
        )
        inserted_details += 1

        # 3. INSERT в offer_channels
        channel_values = []
        for col, ch_name in CHANNEL_COLUMNS.items():
            is_actual = (to_float(row.get(col)) or 0) >= 1
            if is_actual:
                channel_values.append((offer_id, ch_name, True))
        if channel_values:
            execute_values(
                cur,
                "INSERT INTO offer_channels (offer_id, channel, is_actual) VALUES %s "
                "ON CONFLICT (offer_id, channel) DO NOTHING",
                channel_values
            )
            inserted_channels += len(channel_values)

        # 4. INSERT в offer_segments (привязка к поведенческим)
        seg_ids = link_to_behavioral_segments(row, offer_type)
        if seg_ids:
            seg_values = [(offer_id, sid) for sid in seg_ids]
            execute_values(
                cur,
                "INSERT INTO offer_segments (offer_id, segment_id) VALUES %s "
                "ON CONFLICT DO NOTHING",
                seg_values
            )
            inserted_links += len(seg_ids)

    print(f"  ✓ offers: {inserted_offers} офферов")
    print(f"  ✓ offer_details: {inserted_details} записей")
    print(f"  ✓ offer_channels: {inserted_channels} записей")
    print(f"  ✓ offer_segments (привязки): {inserted_links} связей")


# ─────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────

def main():
    if not EXCEL_PATH.exists():
        print(f"ОШИБКА: не найден файл {EXCEL_PATH}")
        sys.exit(1)

    print(f"Подключение к БД {DB_CONFIG['dbname']} на {DB_CONFIG['host']}:{DB_CONFIG['port']}...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except psycopg2.Error as e:
        print(f"ОШИБКА подключения к БД: {e}")
        print("Проверь, что docker-compose up запущен и миграция 10 накачена.")
        sys.exit(1)

    print(f"Открываю {EXCEL_PATH.name}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=False)

    try:
        with conn:
            with conn.cursor() as cur:
                print("\n[1/5] Очистка старых офферов...")
                clear_old_offers(cur)

                print("\n[2/5] Импорт products...")
                products_map = import_products(cur, wb)

                print("\n[3/5] Импорт vas...")
                import_vas(cur, wb)

                print("\n[4/5] Импорт availability_rules...")
                import_availability_rules(cur, wb)

                print("\n[5/5] Импорт офферов...")
                import_offers(cur, wb, products_map)

        print("\n✓ Импорт завершён успешно.")
    except Exception as e:
        print(f"\nОШИБКА при импорте: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        conn.close()


if __name__ == '__main__':
    main()
